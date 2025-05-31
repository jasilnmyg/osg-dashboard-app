import streamlit as st
import pandas as pd
from io import BytesIO
import re
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, PageBreak, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from collections import defaultdict
from reportlab.lib.units import mm
from reportlab.lib.pagesizes import letter
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import streamlit.components.v1 as components

st.set_page_config(
    page_title="OSG DASHBOARD",
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Animated gradient background + glassmorphism + neon glow CSS + tab animations
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Orbitron:wght@700&family=Poppins&display=swap');

:root {
  --dark-bg: #0f111a;
  --panel-bg: rgba(20, 23, 35, 0.75);
  --panel-blur: 12px;
  --neon-blue: #00ffff;
  --neon-pink: #ff00ff;
  --neon-purple: #9d00ff;
  --text-primary: #e0e0ff;
  --text-secondary: #f484fa;
  --glow-shadow: 0 0 8px var(--neon-blue);
  --btn-glow: 0 0 12px var(--neon-pink);
}

/* Animated gradient background */
body, .stApp {
  margin: 0; padding: 0;
  background: linear-gradient(270deg, #090a1a, #1a0a2e, #100a23);
  background-size: 600% 600%;
  animation: gradientShift 20s ease infinite;
  font-family: 'Poppins', sans-serif;
  color: var(--text-primary);
  overflow-x: hidden;
}

@keyframes gradientShift {
  0% {background-position:0% 50%;}
  50% {background-position:100% 50%;}
  100% {background-position:0% 50%;}
}

/* Dashboard header */
.dashboard-header {
  background: var(--panel-bg);
  backdrop-filter: blur(var(--panel-blur));
  border-radius: 1.5rem 1.5rem 0 0;
  padding: 2.2rem 3rem;
  box-shadow: var(--glow-shadow);
  user-select: none;
  text-align: center;
  border: 1px solid var(--neon-blue);
}

.header-title {
  font-family: 'Orbitron', sans-serif;
  font-size: 3rem;
  font-weight: 700;
  color: var(--neon-pink);
  letter-spacing: 0.2em;
  margin-bottom: 0.3rem;
  text-transform: uppercase;
  text-shadow:
    0 0 8px var(--neon-pink),
    0 0 15px var(--neon-pink);
}

.header-subtitle {
  font-size: 1.2rem;
  color: var(--text-secondary);
  font-weight: 500;
  letter-spacing: 0.08em;
}

/* Tabs container */
.stTabs [data-baseweb="tab-list"] {
  display: flex !important;
  justify-content: center;
  gap: 2rem;
  margin-top: 2rem;
  margin-bottom: 2rem;
}

/* Tabs */
.stTabs [data-baseweb="tab"] {
  background: var(--panel-bg);
  backdrop-filter: blur(var(--panel-blur));
  border: 2px solid transparent;
  border-radius: 50px;
  padding: 0.7rem 2.8rem;
  font-weight: 700;
  color: var(--text-secondary);
  font-size: 1.2rem;
  cursor: pointer;
  position: relative;
  transition: color 0.3s ease, border-color 0.3s ease, box-shadow 0.3s ease;
  user-select: none;
  text-transform: uppercase;
  letter-spacing: 0.08em;
  box-shadow: 0 0 6px transparent;
  display: flex;
  align-items: center;
  gap: 0.8rem;
}

/* Neon glow effect on hover */
.stTabs [data-baseweb="tab"]:hover {
  color: var(--neon-pink);
  border-color: var(--neon-pink);
  box-shadow:
    0 0 10px var(--neon-pink),
    0 0 20px var(--neon-pink);
}

/* Active tab neon underline + glow */
.stTabs [aria-selected="true"] {
  color: var(--neon-pink);
  border-color: var(--neon-pink);
  box-shadow:
    0 0 20px var(--neon-pink),
    0 0 30px var(--neon-pink);
}

.stTabs [aria-selected="true"]::after {
  content: "";
  position: absolute;
  bottom: -10px;
  left: 50%;
  transform: translateX(-50%);
  width: 60%;
  height: 3px;
  border-radius: 20px;
  background: linear-gradient(90deg, var(--neon-pink), var(--neon-purple));
  animation: neonPulse 1.8s ease infinite;
  filter: drop-shadow(0 0 5px var(--neon-pink));
}

@keyframes neonPulse {
  0%, 100% {
    opacity: 1;
    filter: drop-shadow(0 0 10px var(--neon-pink));
  }
  50% {
    opacity: 0.5;
    filter: drop-shadow(0 0 5px var(--neon-pink));
  }
}

/* File uploader */
.stFileUpload {
  background: var(--panel-bg) !important;
  border-radius: 1.2rem !important;
  border: 2px dashed var(--neon-blue) !important;
  padding: 3rem 0 !important;
  color: var(--text-secondary) !important;
  font-weight: 600 !important;
  font-size: 1.2rem !important;
  transition: border-color 0.3s ease, box-shadow 0.3s ease !important;
  text-align: center !important;
  user-select: none;
  box-shadow: 0 0 6px transparent !important;
  margin-bottom: 3rem !important;
}

.stFileUpload:hover {
  border-color: var(--neon-pink) !important;
  box-shadow:
    0 0 15px var(--neon-pink) !important;
  color: var(--neon-pink) !important;
  cursor: pointer;
}

/* Button style */
.stButton > button {
  background: linear-gradient(135deg, var(--neon-pink), var(--neon-purple));
  box-shadow:
    0 0 12px var(--neon-pink),
    0 0 18px var(--neon-purple);
  border-radius: 40px;
  padding: 0.8rem 3rem;
  font-weight: 700;
  font-size: 1.15rem;
  color: white;
  transition: all 0.4s ease;
  user-select: none;
}

.stButton > button:hover {
  box-shadow:
    0 0 20px var(--neon-pink),
    0 0 30px var(--neon-purple);
  transform: translateY(-3px);
}

/* Dataframe style */
div[data-testid="stDataFrameContainer"] {
  background: var(--panel-bg);
  border-radius: 1rem;
  padding: 1rem 2rem;
  box-shadow: 0 0 15px rgba(157, 0, 255, 0.2);
  font-family: 'Poppins', sans-serif;
  color: var(--text-primary);
}

/* Scrollbar styling */
::-webkit-scrollbar {
  width: 8px;
  height: 8px;
}
::-webkit-scrollbar-track {
  background: transparent;
}
::-webkit-scrollbar-thumb {
  background: var(--neon-pink);
  border-radius: 20px;
  box-shadow: 0 0 10px var(--neon-pink);
}

/* Icons inside tabs */
.stTabs [data-baseweb="tab"] svg {
  width: 22px;
  height: 22px;
  fill: var(--text-secondary);
  transition: fill 0.3s ease;
}
.stTabs [data-baseweb="tab"]:hover svg,
.stTabs [aria-selected="true"] svg {
  fill: var(--neon-pink);
}
</style>
""", unsafe_allow_html=True)

# Neon glowing icons as SVG for tabs (can also use emojis or images)
tab_icons = {
    "📊 OSG REPORT 1": """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="currentColor"><path d="M13 2h-2v10h2V2zM6 9h2v13H6V9zm10 0h2v13h-2V9z"/></svg>""",
    "📊 OSG REPORT 2": """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="currentColor"><path d="M3 17h2v4H3v-4zm4-6h2v10H7V11zm4-4h2v14h-2V7zm4 6h2v8h-2v-8z"/></svg>""",
    "🔗 Data Mapping": """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="currentColor"><path d="M15.5 14h-.79l-.28-.27A6.471 6.471 0 0016 9.5 6.5 6.5 0 109.5 16c1.61 0 3.09-.59 4.23-1.57l.27.28v.79l5 4.99L20.49 19l-4.99-5zM9.5 14C7.57 14 6 12.43 6 10.5S7.57 7 9.5 7 13 8.57 13 10.5 11.43 14 9.5 14z"/></svg>"""
}

# Streamlit Tabs with icons + neon styles
tab1,tab2,tab3 = st.tabs(list(tab_icons.keys()))



# --------------------------- REPORT 1 TAB ---------------------------
with tab1:
    st.markdown('<h1 class="header">OSG All Store Report</h1>', unsafe_allow_html=True)
    
    with st.container():
        st.markdown("""
        <div class="info-box">
            <strong>Instructions:</strong> Upload the following three files to generate the sales summary report:
            <ul>
                <li><strong>Book1.xlsx</strong> - Sales data with DATE, Store, QUANTITY, AMOUNT fields</li>
                <li><strong>Future store list.xlsx</strong> - Reference list of stores</li>
                <li><strong>RBM and BDM file</strong> - Mapping of Store to RBM and BDM</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    # Date selector at the top
    col1, col2 = st.columns([1, 3])
    with col1:
        report_date = st.date_input("Select report date", value=datetime.today(), key="report1_date")
    
    # File upload section
    with st.container():
        st.markdown('<div class="file-upload-section">', unsafe_allow_html=True)
        book1_file = st.file_uploader(
            "Upload Book1.xlsx (full month sales data)", 
            type=["xlsx"],
            key="book1_uploader"
        )
        store_list_file = st.file_uploader(
            "Upload Future Store List", 
            type=["xlsx"],
            key="store_list_uploader"
        )
        rbm_bdm_file = st.file_uploader(
            "Upload RBM and BDM Mapping File", 
            type=["xlsx"],
            key="rbm_bdm_uploader"
        )
        st.markdown('</div>', unsafe_allow_html=True)

    if book1_file and store_list_file and rbm_bdm_file:
        with st.spinner('Processing data...'):
            # Load and clean data
            book1_df = pd.read_excel(book1_file)
            future_store_df = pd.read_excel(store_list_file)
            rbm_bdm_df = pd.read_excel(rbm_bdm_file)

            # Rename 'Branch' to 'Store'
            book1_df.rename(columns={'Branch': 'Store'}, inplace=True)
            rbm_bdm_df.rename(columns={'Branch': 'Store'}, inplace=True)

            # Parse DATE
            book1_df['DATE'] = pd.to_datetime(book1_df['DATE'], dayfirst=True, errors='coerce')
            book1_df = book1_df.dropna(subset=['DATE'])

            # Use selected report_date for filtering
            today = pd.to_datetime(report_date)

            mtd_df = book1_df[book1_df['DATE'].dt.month == today.month]
            today_df = mtd_df[mtd_df['DATE'].dt.date == today.date()]

            # Aggregate
            today_agg = today_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'FTD Count', 'AMOUNT': 'FTD Amount'})
            mtd_agg = mtd_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'MTD Count', 'AMOUNT': 'MTD Amount'})

            # Merge all unique stores
            all_store_names = pd.Series(pd.concat([future_store_df['Store'], book1_df['Store']]).unique(), name='Store')
            report_df = pd.DataFrame(all_store_names)

            # Merge Today, MTD, and RBM/BDM info
            report_df = report_df.merge(today_agg, on='Store', how='left').merge(mtd_agg, on='Store', how='left')
            report_df[['FTD Count', 'FTD Amount', 'MTD Count', 'MTD Amount']] = report_df[['FTD Count', 'FTD Amount', 'MTD Count', 'MTD Amount']].fillna(0).astype(int)
            report_df = report_df.merge(rbm_bdm_df[['Store', 'RBM', 'BDM']], on='Store', how='left')

            # Sort full report
            report_df = report_df.sort_values('MTD Amount', ascending=False)

            # --- Excel generation ---
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            data_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            zero_qty_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
            total_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            columns_to_use = ['Store', 'FTD Count', 'FTD Amount', 'MTD Count', 'MTD Amount']

            def write_to_sheet(ws, data):
                for r_idx, row in enumerate(dataframe_to_rows(data[columns_to_use], index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                        else:
                            ftd_qty = row[1] if len(row) > 1 else 0
                            mtd_qty = row[3] if len(row) > 3 else 0
                            if ftd_qty == 0 or mtd_qty == 0:
                                cell.fill = zero_qty_fill
                            else:
                                cell.fill = data_fill
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')

                # Add total row
                total_row_idx = ws.max_row + 1
                ws.cell(row=total_row_idx, column=1, value="TOTAL").fill = total_fill
                ws.cell(row=total_row_idx, column=1).font = Font(bold=True)
                ws.cell(row=total_row_idx, column=1).alignment = Alignment(horizontal='center')
                ws.cell(row=total_row_idx, column=1).border = border

                # Leave blank for second column
                ws.cell(row=total_row_idx, column=2, value="").fill = total_fill
                ws.cell(row=total_row_idx, column=2).font = Font(bold=True)
                ws.cell(row=total_row_idx, column=2).alignment = Alignment(horizontal='center')
                ws.cell(row=total_row_idx, column=2).border = border

                # Add totals for FTD Count, Amount, MTD Count, Amount
                for col_idx in range(2, len(columns_to_use) + 1):
                    total_value = data[columns_to_use[col_idx - 1]].sum()
                    cell = ws.cell(row=total_row_idx, column=col_idx, value=int(total_value))
                    cell.fill = total_fill
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')

            wb = Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)

            ws = wb.create_sheet(title="All_Stores")
            write_to_sheet(ws, report_df)

            for rbm in report_df['RBM'].dropna().unique():
                rbm_data = report_df[report_df['RBM'] == rbm].sort_values('MTD Amount', ascending=False)
                ws_rbm = wb.create_sheet(title=rbm[:30])  # Sheet name limit
                write_to_sheet(ws_rbm, rbm_data)

            # Save workbook to BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # --- PDF generation ---
            styles = getSampleStyleSheet()
            base_table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ])

            col_widths = [100*mm/2.54, 70*mm/2.54, 60*mm/2.54, 60*mm/2.54, 60*mm/2.54, 60*mm/2.54]

            rbm_list = report_df['RBM'].dropna().unique()

            pdf_files = []
            for rbm in rbm_list:
                rbm_data = report_df[report_df['RBM'] == rbm].sort_values(by='MTD Amount', ascending=False)
                if rbm_data.empty:
                    continue

                pdf_buffer = BytesIO()
                doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
                elements = []

                elements.append(Paragraph(f"<b><font size=14>{rbm} Report</font></b>", styles['Title']))
                elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%d-%m-%Y')}", styles['Normal']))
                elements.append(Spacer(1, 12))

                table_data = [['Store', 'BDM', 'FTD Count', 'FTD Amount', 'MTD Count', 'MTD Amount']]
                cell_styles = []

                for row_idx, (_, row) in enumerate(rbm_data.iterrows(), start=1):
                    table_row = [
                        row['Store'], row['BDM'],
                        int(row['FTD Count']), int(row['FTD Amount']),
                        int(row['MTD Count']), int(row['MTD Amount'])
                    ]
                    table_data.append(table_row)

                    if row['FTD Count'] == 0:
                        cell_styles.append(('TEXTCOLOR', (2, row_idx), (2, row_idx), colors.red))
                    if row['MTD Count'] == 0:
                        cell_styles.append(('TEXTCOLOR', (4, row_idx), (4, row_idx), colors.red))

                total_row = [
                    'TOTAL', '',
                    int(rbm_data['FTD Count'].sum()), int(rbm_data['FTD Amount'].sum()),
                    int(rbm_data['MTD Count'].sum()), int(rbm_data['MTD Amount'].sum())
                ]
                table_data.append(total_row)
                total_row_idx = len(table_data) - 1
                cell_styles.append(('BACKGROUND', (0, total_row_idx), (-1, total_row_idx), colors.HexColor('#FFD966')))
                cell_styles.append(('FONTNAME', (0, total_row_idx), (-1, total_row_idx), 'Helvetica-Bold'))

                table = Table(table_data, colWidths=col_widths)
                table.setStyle(TableStyle(base_table_style.getCommands() + cell_styles))

                elements.append(table)

                doc.build(elements)
                pdf_buffer.seek(0)

                pdf_files.append((f"{rbm}_Report.pdf", pdf_buffer.read()))

        # Download section
        with st.container():
            st.markdown('<div class="download-section">', unsafe_allow_html=True)
            st.markdown('<h3>Download Reports</h3>', unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            with col1:
                st.download_button(
                    label="📥 Download Excel Report (All Data)",
                    data=excel_buffer,
                    file_name=f"Sales_Report_{report_date.strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Download complete report in Excel format"
                )
            
            with col2:
                st.markdown('<p style="margin-top: 10px;">Individual PDF Reports by RBM:</p>', unsafe_allow_html=True)
                for filename, pdf_data in pdf_files:
                    st.download_button(
                        label=f"📄 {filename.replace('_Report.pdf','')}",
                        data=pdf_data,
                        file_name=filename,
                        mime="application/pdf",
                        key=f"pdf_{filename}"
                    )
            st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.info("ℹ️ Please upload all three required Excel files to generate reports.")

# --------------------------- REPORT 2 TAB ---------------------------
with tab2:
    st.markdown('<h1 class="header">OSG Future Store Report</h1>', unsafe_allow_html=True)
    
    with st.container():
        st.markdown("""
        <div class="info-box">
            <strong>Instructions:</strong> Upload the following files to generate the store summary report:
            <ul>
                <li><strong>Book1.xlsx</strong> - Sales data with Store, QUANTITY, AMOUNT fields</li>
                <li><strong>Future store list.xlsx</strong> - Store master list</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    # File upload section
    with st.container():
        st.markdown('<div class="file-upload-section">', unsafe_allow_html=True)
        book2_file = st.file_uploader(
            "Upload Book1.xlsx", 
            type=["xlsx"],
            key="r2_book1"
        )
        store_list_file = st.file_uploader(
            "Upload Future Store List", 
            type=["xlsx"],
            key="r2_store_list"
        )
        st.markdown('</div>', unsafe_allow_html=True)

    if book2_file and store_list_file:
        with st.spinner('Processing data...'):
            book2_df = pd.read_excel(book2_file)
            future_df = pd.read_excel(store_list_file)

            book2_df.rename(columns={'Branch': 'Store'}, inplace=True)
            agg = book2_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'})

            all_stores = pd.DataFrame(pd.concat([future_df['Store'], agg['Store']]).unique(), columns=['Store'])
            merged = all_stores.merge(agg, on='Store', how='left')
            merged['QUANTITY'] = merged['QUANTITY'].fillna(0).astype(int)
            merged['AMOUNT'] = merged['AMOUNT'].fillna(0).astype(int)

            merged = merged.sort_values(by='AMOUNT', ascending=False).reset_index(drop=True)
            total = pd.DataFrame([{
                'Store': 'TOTAL',
                'QUANTITY': merged['QUANTITY'].sum(),
                'AMOUNT': merged['AMOUNT'].sum()
            }])
            final_df = pd.concat([merged, total], ignore_index=True)
            final_df.rename(columns={'Store': 'Branch'}, inplace=True)

            def generate_report2_excel(df):
                wb = Workbook()
                ws = wb.active
                ws.title = "Store Report"

                header_fill = PatternFill("solid", fgColor="4F81BD")
                data_fill = PatternFill("solid", fgColor="DCE6F1")
                red_fill = PatternFill("solid", fgColor="F4CCCC")
                total_fill = PatternFill("solid", fgColor="FFD966")
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
                bold_font = Font(bold=True)
                header_font = Font(bold=True, color="FFFFFF")

                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                        elif df.loc[r_idx - 2, 'Branch'] == 'TOTAL':
                            cell.fill = total_fill
                            cell.font = bold_font
                        elif df.loc[r_idx - 2, 'AMOUNT'] <= 0:
                            cell.fill = red_fill
                        else:
                            cell.fill = data_fill
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')

                for column_cells in ws.columns:
                    length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                    ws.column_dimensions[column_cells[0].column_letter].width = length + 2

                buf = BytesIO()
                wb.save(buf)
                buf.seek(0)
                return buf

            excel_buf2 = generate_report2_excel(final_df)
        
        # Download section
        with st.container():
            st.markdown('<div class="download-section">', unsafe_allow_html=True)
            st.download_button(
                label="📥 Download Store Summary Report",
                data=excel_buf2,
                file_name="Store_Summary_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Download store summary report in Excel format"
            )
            st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.info("ℹ️ Please upload both required files to generate the store summary report.")

# --------------------------- REPORT 3 TAB ---------------------------
with tab3:
    st.markdown('<h1 class="header">OSG & Product Data Mapping</h1>', unsafe_allow_html=True)
    
    with st.container():
        st.markdown("""
        <div class="info-box">
            <strong>Instructions:</strong> Upload the following files to map OSG and product data:
            <ul>
                <li><strong>OSG File</strong> - Contains warranty and protection plan data</li>
                <li><strong>PRODUCT File</strong> - Contains product information including models, categories, and IMEIs</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    # File upload section
    with st.container():
        st.markdown('<div class="file-upload-section">', unsafe_allow_html=True)
        osg_file = st.file_uploader(
            "Upload OSG File", 
            type=["xlsx"],
            key="osg_mapping"
        )
        product_file = st.file_uploader(
            "Upload PRODUCT File", 
            type=["xlsx"],
            key="product_mapping"
        )
        st.markdown('</div>', unsafe_allow_html=True)

    if osg_file and product_file:
        with st.spinner('Mapping data...'):
            osg_df = pd.read_excel(osg_file)
            product_df = pd.read_excel(product_file)

            # SKU Mapping
            sku_category_mapping = {
                "Warranty : Water Cooler/Dispencer/Geyser/RoomCooler/Heater": [
                    "COOLER", "DISPENCER", "GEYSER", "ROOM COOLER", "HEATER", "WATER HEATER", "WATER DISPENSER"
                ],
                "Warranty : Fan/Mixr/IrnBox/Kettle/OTG/Grmr/Geysr/Steamr/Inductn": [
                    "FAN", "MIXER", "IRON BOX", "KETTLE", "OTG", "GROOMING KIT", "GEYSER", "STEAMER", "INDUCTION",
                    "CEILING FAN", "TOWER FAN", "PEDESTAL FAN", "INDUCTION COOKER", "ELECTRIC KETTLE", "WALL FAN", "MIXER GRINDER", "CELLING FAN"
                ],
                "AC : EWP : Warranty : AC": ["AC", "AIR CONDITIONER", "AC INDOOR"],
                "HAEW : Warranty : Air Purifier/WaterPurifier": ["AIR PURIFIER", "WATER PURIFIER"],
                "HAEW : Warranty : Dryer/MW/DishW": ["DRYER", "MICROWAVE OVEN", "DISH WASHER", "MICROWAVE OVEN-CONV"],
                "HAEW : Warranty : Ref/WM": [
                    "REFRIGERATOR", "WASHING MACHINE", "WASHING MACHINE-TL", "REFRIGERATOR-DC",
                    "WASHING MACHINE-FL", "WASHING MACHINE-SA", "REF", "REFRIGERATOR-CBU", "REFRIGERATOR-FF", "WM"
                ],
                "HAEW : Warranty : TV": ["TV", "TV 28 %", "TV 18 %"],
                "TV : TTC : Warranty and Protection : TV": ["TV", "TV 28 %", "TV 18 %"],
                "TV : Spill and Drop Protection": ["TV", "TV 28 %", "TV 18 %"],
                "HAEW : Warranty :Chop/Blend/Toast/Air Fryer/Food Processr/JMG/Induction": [
                    "CHOPPER", "BLENDER", "TOASTER", "AIR FRYER", "FOOD PROCESSOR", "JUICER", "INDUCTION COOKER"
                ],
                "HAEW : Warranty : HOB and Chimney": ["HOB", "CHIMNEY"],
                "HAEW : Warranty : HT/SoundBar/AudioSystems/PortableSpkr": [
                    "HOME THEATRE", "AUDIO SYSTEM", "SPEAKER", "SOUND BAR", "PARTY SPEAKER"
                ],
                "HAEW : Warranty : Vacuum Cleaner/Fans/Groom&HairCare/Massager/Iron": [
                    "VACUUM CLEANER", "FAN", "MASSAGER", "IRON BOX", "CEILING FAN", "TOWER FAN", "PEDESTAL FAN", "WALL FAN", "ROBO VACCUM CLEANER"
                ],
                "AC AMC": ["AC", "AC INDOOR"]
            }

            product_df['Category'] = product_df['Category'].str.upper().fillna('')
            product_df['Model'] = product_df['Model'].fillna('')
            product_df['Customer Mobile'] = product_df['Customer Mobile'].astype(str)
            product_df['Invoice Number'] = product_df['Invoice Number'].astype(str)
            product_df['Item Rate'] = pd.to_numeric(product_df['Item Rate'], errors='coerce')
            product_df['IMEI'] = product_df['IMEI'].astype(str).fillna('')
            product_df['Brand'] = product_df['Brand'].fillna('')
            osg_df['Customer Mobile'] = osg_df['Customer Mobile'].astype(str)

            def extract_price_slab(text):
                match = re.search(r"Slab\s*:\s*(\d+)K-(\d+)K", str(text))
                if match:
                    return int(match.group(1)) * 1000, int(match.group(2)) * 1000
                return None, None

            def get_model(row):
                mobile = row['Customer Mobile']
                retailer_sku = str(row['Retailer SKU'])
                invoice = str(row.get('Invoice Number', ''))
                user_products = product_df[product_df['Customer Mobile'] == mobile]

                if user_products.empty:
                    return ''
                unique_models = user_products['Model'].dropna().unique()
                if len(unique_models) == 1:
                    return unique_models[0]

                mapped_keywords = []
                for sku_key, keywords in sku_category_mapping.items():
                    if sku_key in retailer_sku:
                        mapped_keywords = [kw.lower() for kw in keywords]
                        break

                filtered = user_products[user_products['Category'].str.lower().isin(mapped_keywords)]
                if filtered['Model'].nunique() == 1:
                    return filtered['Model'].iloc[0]

                slab_min, slab_max = extract_price_slab(retailer_sku)
                if slab_min and slab_max:
                    slab_filtered = filtered[(filtered['Item Rate'] >= slab_min) & (filtered['Item Rate'] <= slab_max)]
                    if slab_filtered['Model'].nunique() == 1:
                        return slab_filtered['Model'].iloc[0]
                    invoice_filtered = slab_filtered[slab_filtered['Invoice Number'].astype(str) == invoice]
                    if invoice_filtered['Model'].nunique() == 1:
                        return invoice_filtered['Model'].iloc[0]

                return ''

            osg_df['Model'] = osg_df.apply(get_model, axis=1)
            category_brand_df = product_df[['Customer Mobile', 'Model', 'Category', 'Brand']].drop_duplicates()
            osg_df = osg_df.merge(category_brand_df, on=['Customer Mobile', 'Model'], how='left')

            invoice_pool = defaultdict(list)
            itemrate_pool = defaultdict(list)
            imei_pool = defaultdict(list)

            for _, row in product_df.iterrows():
                key = (row['Customer Mobile'], row['Model'])
                invoice_pool[key].append(row['Invoice Number'])
                itemrate_pool[key].append(row['Item Rate'])
                imei_pool[key].append(row['IMEI'])

            invoice_usage_counter = defaultdict(int)
            itemrate_usage_counter = defaultdict(int)
            imei_usage_counter = defaultdict(int)

            def assign_from_pool(row, pool, counter_dict):
                key = (row['Customer Mobile'], row['Model'])
                values = pool.get(key, [])
                index = counter_dict[key]
                if index < len(values):
                    counter_dict[key] += 1
                    return values[index]
                return ''

            osg_df['Product Invoice Number'] = osg_df.apply(lambda row: assign_from_pool(row, invoice_pool, invoice_usage_counter), axis=1)
            osg_df['Item Rate'] = osg_df.apply(lambda row: assign_from_pool(row, itemrate_pool, itemrate_usage_counter), axis=1)
            osg_df['IMEI'] = osg_df.apply(lambda row: assign_from_pool(row, imei_pool, imei_usage_counter), axis=1)
            osg_df['Store Code'] = osg_df['Product Invoice Number'].astype(str).apply(
                lambda x: re.search(r'\b([A-Z]{2,})\b', x).group(1) if re.search(r'\b([A-Z]{2,})\b', x) else ''
            )

            def extract_warranty_duration(sku):
                sku = str(sku)
                match = re.search(r'Dur\s*:\s*(\d+)\+(\d+)', sku)
                if match:
                    return int(match.group(1)), int(match.group(2))
                match = re.search(r'(\d+)\+(\d+)\s*SDP-(\d+)', sku)
                if match:
                    return int(match.group(1)), f"{match.group(3)}P+{match.group(2)}W"
                match = re.search(r'Dur\s*:\s*(\d+)', sku)
                if match:
                    return 1, int(match.group(1))
                match = re.search(r'(\d+)\+(\d+)', sku)
                if match:
                    return int(match.group(1)), int(match.group(2))
                return '', ''

            osg_df[['Manufacturer Warranty', 'Duration (Year)']] = osg_df['Retailer SKU'].apply(
                lambda sku: pd.Series(extract_warranty_duration(sku))
            )
            
            def highlight_row(row):
                missing_fields = pd.isna(row.get('Model')) or str(row.get('Model')).strip() == ''
                missing_fields |= pd.isna(row.get('IMEI')) or str(row.get('IMEI')).strip() == ''
                try:
                    if float(row.get('Plan Price', 0)) < 0:
                        missing_fields |= True
                except:
                    missing_fields |= True
                return ['background-color: lightblue'] * len(row) if missing_fields else [''] * len(row)
            
            final_columns = [
                'Customer Mobile', 'Date', 'Invoice Number','Product Invoice Number', 'Customer Name', 'Store Code', 'Branch', 'Region',
                'IMEI', 'Category', 'Brand', 'Quantity', 'Item Code', 'Model', 'Plan Type', 'EWS QTY', 'Item Rate',
                'Plan Price', 'Sold Price', 'Email', 'Product Count', 'Manufacturer Warranty', 'Retailer SKU', 'OnsiteGo SKU',
                'Duration (Year)', 'Total Coverage', 'Comment', 'Return Flag', 'Return against invoice No.',
                'Primary Invoice No.'
            ]

            for col in final_columns:
                if col not in osg_df.columns:
                    osg_df[col] = ''
            osg_df['Quantity'] = 1
            osg_df['EWS QTY'] = 1
            osg_df = osg_df[final_columns]
            
            st.markdown("""
            <div class="success-box">
                <strong>✅ Data Mapping Completed Successfully</strong>
                <p>The OSG and product data has been successfully mapped. You can now download the report.</p>
            </div>
            """, unsafe_allow_html=True)
            
            @st.cache_data
            def convert_df(df):
               output = io.BytesIO()
               styled_df = df.style.apply(highlight_row, axis=1)
               with pd.ExcelWriter(output, engine='openpyxl') as writer:
                styled_df.to_excel(writer, index=False)
               output.seek(0)
               return output
            
            excel_data = convert_df(osg_df)
        
        # Download section
        with st.container():
            st.markdown('<div class="download-section">', unsafe_allow_html=True)
            st.download_button(
                label="📥 Download Mapped Data Report",
                data=excel_data,
                file_name="OSG_Product_Mapping_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Download the mapped OSG and product data in Excel format"
            )
            st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.info("ℹ️ Please upload both required files to perform data mapping.")
